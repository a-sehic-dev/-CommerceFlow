"""Professional openpyxl workbook formatting for CommerceFlow enterprise exports."""

from __future__ import annotations

from datetime import datetime
from io import BytesIO
from typing import Any

import pandas as pd
from app.utils.app_timezone import (
    APP_TZ_NAME,
    filename_timestamp,
    format_display,
    naive_local_now,
)
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.utils.excel_polish import (
    TableBounds,
    apply_sheet_conditional_formatting,
    apply_workbook_tab_colors,
    protect_executive_sheet,
    write_executive_charts,
)

ENGINE_VERSION = "CommerceFlow 1.0.0 · Sedin Sehic"
EXEC_COLS = 12  # A–L presentation canvas

# ── Theme ────────────────────────────────────────────────────────────────────
TITLE_FONT = Font(name="Calibri", bold=True, size=18, color="0F172A")
SUBTITLE_FONT = Font(name="Calibri", size=11, color="475569", italic=True)
META_LABEL_FONT = Font(name="Calibri", bold=True, size=10, color="334155")
META_VALUE_FONT = Font(name="Calibri", size=10, color="1E293B")
SECTION_FONT = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
BODY_FONT = Font(name="Calibri", size=10, color="1E293B")

HERO_FILL = PatternFill("solid", fgColor="0F172A")
HERO_TITLE_FONT = Font(name="Calibri", bold=True, size=24, color="FFFFFF")
HERO_LINE_FONT = Font(name="Calibri", size=11, color="CBD5E1")
HERO_ACCENT_FONT = Font(name="Calibri", size=10, color="94A3B8")

KPI_SECTION_FONT = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
KPI_LABEL_FONT = Font(name="Calibri", bold=True, size=9, color="64748B")
KPI_VALUE_FONT = Font(name="Calibri", bold=True, size=22, color="0F172A")
KPI_RISK_VALUE_FONT = Font(name="Calibri", bold=True, size=22, color="B91C1C")
KPI_CARD_FILL = PatternFill("solid", fgColor="FFFFFF")
KPI_CARD_ACCENT = PatternFill("solid", fgColor="EEF2FF")
KPI_CARD_RISK_FILL = PatternFill("solid", fgColor="FEF2F2")
KPI_SCORE_MID_FONT = Font(name="Calibri", bold=True, size=22, color="C2410C")
KPI_SCORE_OK_FONT = Font(name="Calibri", bold=True, size=22, color="0F172A")

AUDIT_SECTION_FONT = Font(name="Calibri", bold=True, size=10, color="475569")
AUDIT_HEADER_FILL = PatternFill("solid", fgColor="E2E8F0")
AUDIT_HEADER_FONT = Font(name="Calibri", bold=True, size=9, color="475569")
AUDIT_BODY_FONT = Font(name="Calibri", size=9, color="334155")
AUDIT_ALT_FILL = PatternFill("solid", fgColor="F8FAFC")
AUDIT_BODY_FILL = PatternFill("solid", fgColor="FFFFFF")

SECTION_FILL = PatternFill("solid", fgColor="0F172A")
HEADER_FILL = PatternFill("solid", fgColor="1E293B")
ALT_FILL = PatternFill("solid", fgColor="F8FAFC")
META_FILL = PatternFill("solid", fgColor="F1F5F9")

CARD_BORDER = Border(
    left=Side(style="medium", color="CBD5E1"),
    right=Side(style="medium", color="CBD5E1"),
    top=Side(style="medium", color="CBD5E1"),
    bottom=Side(style="medium", color="CBD5E1"),
)

THIN_BORDER = Border(
    left=Side(style="thin", color="CBD5E1"),
    right=Side(style="thin", color="CBD5E1"),
    top=Side(style="thin", color="CBD5E1"),
    bottom=Side(style="thin", color="CBD5E1"),
)

TOP_ALIGN = Alignment(horizontal="left", vertical="top", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="top")
CENTER = Alignment(horizontal="center", vertical="center")
BANNER_TITLE_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
BANNER_SUBTITLE_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
BANNER_META_ALIGN = Alignment(horizontal="left", vertical="center", indent=1)

BANNER_TITLE_ROW_HEIGHT = 38
BANNER_SUBTITLE_ROW_HEIGHT = 24
BANNER_GENERATED_ROW_HEIGHT = 20
SECTION_BAR_ROW_HEIGHT = 30

CURRENCY_FMT = '"$"#,##0.00'
PCT_FMT = "0.0%"
COUNT_FMT = "#,##0"
SCORE_FMT = "0.0"
GENERAL_FMT = "General"

KPI_SPECS: list[tuple[str, str, str]] = [
    ("Total Revenue", "total_revenue", "currency"),
    ("Gross Margin %", "gross_margin_pct", "percent"),
    ("Inventory Efficiency %", "inventory_efficiency", "percent"),
    ("Operational Risk Score", "operational_risk_score", "score"),
    ("Avg Order Value", "avg_order_value", "currency"),
    ("Profit Leakage Estimate", "profit_leakage_estimate", "exposure_currency"),
    ("Total Alerts", "active_alerts", "count"),
    ("Active Products", "product_count", "count"),
    ("Dead Inventory Value", "dead_inventory_value", "currency"),
    ("Total Orders", "total_orders", "count"),
]

METRIC_KEY_LABELS: dict[str, str] = {
    "total_revenue": "Total Revenue",
    "total_orders": "Total Orders",
    "avg_order_value": "Avg Order Value",
    "gross_margin_pct": "Gross Margin %",
    "inventory_efficiency": "Inventory Efficiency %",
    "operational_risk_score": "Operational Risk Score",
    "product_count": "Active Products",
    "dead_inventory_value": "Dead Inventory Value",
    "profit_leakage_estimate": "Profit Leakage Estimate",
    "active_alerts": "Total Alerts",
}


def build_enterprise_workbook(
    *,
    metrics: dict,
    metric_traces: dict,
    analysis: dict,
    products_rows: list[dict],
    inventory_rows: list[dict],
    alerts_rows: list[dict],
    metadata: dict[str, Any] | None = None,
) -> bytes:
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    meta = metadata or {}
    generated_at = meta.get("generated_at") or naive_local_now()

    chart_data = meta.get("chart_data") or {}
    _write_executive_summary_sheet(wb, metrics, metric_traces, meta, generated_at, chart_data)

    intel = analysis.get("product_intelligence", {})
    product_records: list[dict] = []
    for label, items in [
        ("Top Seller", intel.get("top_sellers", [])),
        ("Underperformer", intel.get("worst_performers", [])),
    ]:
        for item in items:
            row = dict(item)
            row["performance_tier"] = label
            product_records.append(row)

    if products_rows:
        for p in products_rows:
            if not any(r.get("sku") == p.get("sku") for r in product_records):
                product_records.append(p)

    prod_df = _products_dataframe(product_records)
    _write_report_sheet(
        wb,
        "Products Intelligence",
        title="Products Intelligence",
        subtitle="Catalog performance, health scores, and revenue contribution",
        generated_at=generated_at,
        df=prod_df,
    )

    inv_df = _inventory_dataframe(analysis, inventory_rows)
    _write_report_sheet(
        wb,
        "Inventory Health",
        title="Inventory Health",
        subtitle="Stock risk, reorder signals, and inventory health scores",
        generated_at=generated_at,
        df=inv_df,
    )

    alerts_df = _alerts_dataframe(alerts_rows)
    _write_report_sheet(
        wb,
        "Alerts Center",
        title="Alerts Center",
        subtitle="Operational alerts ranked by severity and recency",
        generated_at=generated_at,
        df=alerts_df,
    )

    leak_df = _profit_leakage_dataframe(analysis)
    _write_report_sheet(
        wb,
        "Profit Leakage",
        title="Profit Leakage",
        subtitle="Recoverable margin issues with estimated business impact",
        generated_at=generated_at,
        df=leak_df,
    )

    apply_workbook_tab_colors(wb)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def write_dataframe_sheet(
    wb: Workbook,
    title: str,
    df: pd.DataFrame,
    *,
    start_row: int = 1,
    max_rows: int | None = None,
) -> None:
    """Backward-compatible simple sheet writer (summary exports)."""
    _write_report_sheet(
        wb,
        title[:31],
        title=title,
        subtitle="CommerceFlow export",
        generated_at=naive_local_now(),
        df=df,
        start_row=start_row,
        max_rows=max_rows,
    )


def write_key_value_sheet(wb: Workbook, title: str, rows: list[dict[str, Any]]) -> None:
    df = pd.DataFrame(rows)
    if "metric" in df.columns and "value" in df.columns:
        df = df.rename(columns={"metric": "Metric", "value": "Value"})
    write_dataframe_sheet(wb, title, df)


def enterprise_filename() -> str:
    ts = filename_timestamp()
    return f"commerceflow_report_{ts}.xlsx"


# ── Executive Summary ─────────────────────────────────────────────────────────


def _write_executive_summary_sheet(
    wb: Workbook,
    metrics: dict,
    metric_traces: dict,
    meta: dict[str, Any],
    generated_at: datetime,
    chart_data: dict[str, Any] | None = None,
) -> None:
    ws = wb.create_sheet(title="Executive Summary", index=0)
    _init_executive_canvas(ws)

    row = _write_executive_hero(ws, meta, generated_at)
    row += 1
    row = _write_kpi_section_bar(ws, row, "Executive KPIs")
    row = _write_kpi_card_grid(ws, row, metrics)
    row = _write_spacer_row(ws, row, height=12)
    row = write_executive_charts(ws, row, chart_data or {})
    row = _write_spacer_row(ws, row, height=14)
    audit_start = row
    row = _write_audit_section(ws, row, metric_traces)
    protect_executive_sheet(ws, unlock_from_row=audit_start)

    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 95


def _init_executive_canvas(ws: Worksheet) -> None:
    for col in range(1, EXEC_COLS + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18
    ws.sheet_properties.pageSetUpPr.fitToPage = True


def _write_executive_hero(ws: Worksheet, meta: dict[str, Any], generated_at: datetime) -> int:
    sales_ds = meta.get("sales_dataset") or "Not selected"
    products_ds = meta.get("products_dataset") or "Not selected"
    inventory_ds = meta.get("inventory_dataset") or "Not selected"
    sales_rows = _format_count(meta.get("sales_rows"))
    products_rows = _format_count(meta.get("products_rows"))
    inventory_rows = _format_count(meta.get("inventory_rows"))

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=EXEC_COLS)
    title_cell = ws.cell(row=1, column=1, value="CommerceFlow Executive Intelligence Report")
    title_cell.fill = HERO_FILL
    title_cell.font = HERO_TITLE_FONT
    title_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=2)
    ws.row_dimensions[1].height = 40
    ws.row_dimensions[2].height = 8

    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=EXEC_COLS)
    ts_cell = ws.cell(
        row=3,
        column=1,
        value=f"Generated {_format_datetime(meta.get('generated_at') or generated_at)} ({APP_TZ_NAME})",
    )
    ts_cell.fill = HERO_FILL
    ts_cell.font = HERO_LINE_FONT
    ts_cell.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[3].height = 20

    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=EXEC_COLS)
    id_cell = ws.cell(
        row=4,
        column=1,
        value=(
            f"Analysis ID  {meta.get('analysis_id') or '—'}    ·    "
            f"Engine  {meta.get('engine_version') or ENGINE_VERSION}"
        ),
    )
    id_cell.fill = HERO_FILL
    id_cell.font = HERO_ACCENT_FONT
    id_cell.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[4].height = 18

    ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=EXEC_COLS)
    ds_cell = ws.cell(
        row=5,
        column=1,
        value=f"Datasets  Sales: {sales_ds}  |  Products: {products_ds}  |  Inventory: {inventory_ds}",
    )
    ds_cell.fill = HERO_FILL
    ds_cell.font = HERO_ACCENT_FONT
    ds_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=2)
    ws.row_dimensions[5].height = 32

    ws.merge_cells(start_row=6, start_column=1, end_row=6, end_column=EXEC_COLS)
    rows_cell = ws.cell(
        row=6,
        column=1,
        value=f"Volume  Sales {sales_rows} rows  ·  Products {products_rows} rows  ·  Inventory {inventory_rows} rows",
    )
    rows_cell.fill = HERO_FILL
    rows_cell.font = HERO_ACCENT_FONT
    rows_cell.alignment = Alignment(horizontal="left", vertical="top", indent=2)
    ws.row_dimensions[6].height = 20

    for r in range(1, 7):
        for c in range(1, EXEC_COLS + 1):
            ws.cell(row=r, column=c).fill = HERO_FILL

    return 8


def _write_spacer_row(ws: Worksheet, row: int, *, height: int = 12) -> int:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=EXEC_COLS)
    spacer = ws.cell(row=row, column=1)
    spacer.fill = PatternFill("solid", fgColor="FFFFFF")
    ws.row_dimensions[row].height = height
    return row + 1


def _write_kpi_section_bar(ws: Worksheet, row: int, title: str) -> int:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=EXEC_COLS)
    cell = ws.cell(row=row, column=1, value=title.upper())
    cell.fill = SECTION_FILL
    cell.font = KPI_SECTION_FONT
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=2)
    ws.row_dimensions[row].height = SECTION_BAR_ROW_HEIGHT
    return row + 1


def _write_kpi_card_grid(ws: Worksheet, start_row: int, metrics: dict) -> int:
    """Five KPI cards per row across the full sheet width (presentation layout)."""
    cards_per_row = 5
    card_height = 3
    gap_rows = 2

    # Five cards per row: A-B, C-D, E-F, G-H, I-L (last card spans 4 columns)
    row_layouts: list[tuple[int, int]] = [
        (1, 2),
        (3, 2),
        (5, 2),
        (7, 2),
        (9, 4),
    ]

    row = start_row
    for idx, (label, key, kind) in enumerate(KPI_SPECS):
        pos = idx % cards_per_row
        band = idx // cards_per_row
        if pos == 0 and band > 0:
            row += card_height + gap_rows

        left_col, width = row_layouts[pos]
        _write_kpi_card(
            ws,
            top_row=row,
            left_col=left_col,
            width=width,
            height=card_height,
            label=label,
            value=metrics.get(key),
            kind=kind,
            metric_key=key,
            accent=(band == 0 and pos % 2 == 0),
        )

    return row + card_height


def _write_kpi_card(
    ws: Worksheet,
    *,
    top_row: int,
    left_col: int,
    width: int,
    height: int,
    label: str,
    value: Any,
    kind: str,
    metric_key: str = "",
    accent: bool = False,
) -> None:
    right_col = left_col + width - 1
    bottom_row = top_row + height - 1
    is_leakage = metric_key == "profit_leakage_estimate" or kind == "exposure_currency"
    is_risk_score = metric_key == "operational_risk_score"
    fill = KPI_CARD_RISK_FILL if is_leakage else (KPI_CARD_ACCENT if accent else KPI_CARD_FILL)

    for r in range(top_row, bottom_row + 1):
        ws.row_dimensions[r].height = 28 if r == top_row else (44 if r == top_row + 1 else 10)
        for c in range(left_col, right_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = fill
            cell.border = CARD_BORDER

    ws.merge_cells(
        start_row=top_row,
        start_column=left_col,
        end_row=top_row,
        end_column=right_col,
    )
    label_cell = ws.cell(row=top_row, column=left_col, value=label.upper())
    label_cell.font = KPI_LABEL_FONT
    label_cell.fill = fill
    label_cell.alignment = Alignment(horizontal="left", vertical="bottom", indent=1)
    label_cell.border = CARD_BORDER

    ws.merge_cells(
        start_row=top_row + 1,
        start_column=left_col,
        end_row=bottom_row,
        end_column=right_col,
    )
    value_cell = ws.cell(row=top_row + 1, column=left_col)
    value_cell.fill = fill
    if is_leakage:
        value_cell.font = KPI_RISK_VALUE_FONT
    elif is_risk_score and value is not None:
        try:
            risk_val = float(value)
            value_cell.font = KPI_SCORE_MID_FONT if risk_val >= 82 else KPI_SCORE_OK_FONT
        except (TypeError, ValueError):
            value_cell.font = KPI_VALUE_FONT
    else:
        value_cell.font = KPI_VALUE_FONT
    value_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    value_cell.border = CARD_BORDER
    _apply_typed_value(value_cell, value, kind)


def _write_audit_section(ws: Worksheet, start_row: int, metric_traces: dict) -> int:
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=EXEC_COLS)
    title = ws.cell(row=start_row, column=1, value="Calculation Trace / Audit Log")
    title.font = AUDIT_SECTION_FONT
    title.fill = PatternFill("solid", fgColor="F1F5F9")
    title.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=2)
    ws.row_dimensions[start_row].height = SECTION_BAR_ROW_HEIGHT
    row = start_row + 1

    headers = ["Metric Key", "Formula Used", "Dataset Source", "Rows Processed", "Notes"]
    col_layout = [1, 3, 5, 7, 9]
    header_row = row
    for header, col in zip(headers, col_layout):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.fill = AUDIT_HEADER_FILL
        cell.font = AUDIT_HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = LEFT
    ws.merge_cells(start_row=header_row, start_column=3, end_row=header_row, end_column=4)
    ws.merge_cells(start_row=header_row, start_column=9, end_row=header_row, end_column=EXEC_COLS)
    for c in range(1, EXEC_COLS + 1):
        hc = ws.cell(row=header_row, column=c)
        hc.fill = AUDIT_HEADER_FILL
        hc.border = THIN_BORDER
    ws.row_dimensions[header_row].height = 18
    row += 1

    trace_rows: list[list[Any]] = []
    for key, trace in metric_traces.items():
        if not isinstance(trace, dict):
            continue
        trace_rows.append([
            METRIC_KEY_LABELS.get(key, key.replace("_", " ").title()),
            trace.get("formula") or "",
            trace.get("dataset") or "",
            trace.get("row_count"),
            _trace_notes(trace),
        ])
    if not trace_rows:
        trace_rows = [["—", "No trace metadata available", "", "", ""]]

    for offset, values in enumerate(trace_rows):
        excel_row = row + offset
        body_fill = AUDIT_ALT_FILL if offset % 2 else AUDIT_BODY_FILL
        ws.row_dimensions[excel_row].height = 36

        key_cell = ws.cell(row=excel_row, column=1, value=values[0])
        key_cell.font = AUDIT_BODY_FONT
        key_cell.fill = body_fill
        key_cell.border = THIN_BORDER
        key_cell.alignment = TOP_ALIGN

        ws.merge_cells(start_row=excel_row, start_column=3, end_row=excel_row, end_column=4)
        formula_cell = ws.cell(row=excel_row, column=3, value=values[1])
        formula_cell.font = AUDIT_BODY_FONT
        formula_cell.fill = body_fill
        formula_cell.border = THIN_BORDER
        formula_cell.alignment = TOP_ALIGN

        ds_cell = ws.cell(row=excel_row, column=5, value=values[2])
        ds_cell.font = AUDIT_BODY_FONT
        ds_cell.fill = body_fill
        ds_cell.border = THIN_BORDER
        ds_cell.alignment = TOP_ALIGN

        rows_cell = ws.cell(row=excel_row, column=7, value=values[3])
        rows_cell.font = AUDIT_BODY_FONT
        rows_cell.fill = body_fill
        rows_cell.border = THIN_BORDER
        rows_cell.alignment = RIGHT
        if values[3] not in (None, ""):
            try:
                rows_cell.value = int(float(values[3]))
                rows_cell.number_format = COUNT_FMT
            except (TypeError, ValueError):
                rows_cell.value = _clean_value(values[3])

        ws.merge_cells(
            start_row=excel_row,
            start_column=9,
            end_row=excel_row,
            end_column=EXEC_COLS,
        )
        notes_cell = ws.cell(row=excel_row, column=9, value=values[4])
        notes_cell.font = AUDIT_BODY_FONT
        notes_cell.fill = body_fill
        notes_cell.border = THIN_BORDER
        notes_cell.alignment = TOP_ALIGN

    last_row = row + len(trace_rows) - 1
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(EXEC_COLS)}{last_row}"
    ws.freeze_panes = ws.cell(row=row, column=1)
    return last_row + 2


# ── Standard report sheets ────────────────────────────────────────────────────


def _write_report_sheet(
    wb: Workbook,
    sheet_name: str,
    *,
    title: str,
    subtitle: str,
    generated_at: datetime,
    df: pd.DataFrame,
    start_row: int = 1,
    max_rows: int | None = None,
) -> None:
    from app.config import get_settings

    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])
    ws = wb.create_sheet(title=sheet_name[:31])

    if max_rows is None:
        max_rows = get_settings().export_max_rows_per_sheet
    if len(df) > max_rows:
        df = df.head(max_rows)

    row = start_row
    col_count = max(len(df.columns), 4) if not df.empty else 4
    row = _write_sheet_banner(
        ws,
        row,
        title=title,
        subtitle=subtitle,
        generated_at=generated_at,
        merge_cols=col_count,
    )
    row += 1

    if df.empty:
        ws.cell(row=row, column=1, value="No data available for this section.")
        ws.cell(row=row, column=1).font = Font(italic=True, color="64748B")
        ws.freeze_panes = ws.cell(row=1, column=1)
        return

    headers = [str(c) for c in df.columns]
    rows = [
        [_clean_value(v) for v in record]
        for record in df.itertuples(index=False, name=None)
    ]
    formats = [_column_format(h) for h in headers]
    wrap = {i + 1 for i, h in enumerate(headers) if _should_wrap(h)}

    bounds = _write_formatted_table(
        ws,
        row,
        headers,
        rows,
        column_formats=formats,
        freeze=True,
        table_name=f"T_{sheet_name.replace(' ', '_')[:18]}",
        wrap_columns=wrap,
    )
    apply_sheet_conditional_formatting(ws, bounds, sheet_name)
    _auto_width(ws, headers=headers, min_width=12, max_width=52)


# ── Layout primitives ─────────────────────────────────────────────────────────


def _write_sheet_banner(
    ws: Worksheet,
    start_row: int,
    *,
    title: str,
    subtitle: str,
    generated_at: datetime,
    merge_cols: int,
) -> int:
    merge_cols = max(merge_cols, 2)
    end_col = get_column_letter(merge_cols)

    title_row = start_row
    ws.merge_cells(f"A{title_row}:{end_col}{title_row}")
    c = ws.cell(row=title_row, column=1, value=title)
    c.font = TITLE_FONT
    c.alignment = BANNER_TITLE_ALIGN
    ws.row_dimensions[title_row].height = BANNER_TITLE_ROW_HEIGHT

    subtitle_row = start_row + 1
    ws.merge_cells(f"A{subtitle_row}:{end_col}{subtitle_row}")
    c = ws.cell(row=subtitle_row, column=1, value=subtitle)
    c.font = SUBTITLE_FONT
    c.alignment = BANNER_SUBTITLE_ALIGN
    ws.row_dimensions[subtitle_row].height = BANNER_SUBTITLE_ROW_HEIGHT

    generated_row = start_row + 2
    ws.merge_cells(f"A{generated_row}:{end_col}{generated_row}")
    c = ws.cell(
        row=generated_row,
        column=1,
        value=f"Generated {_format_datetime(generated_at)} ({APP_TZ_NAME})",
    )
    c.font = Font(name="Calibri", size=9, color="64748B")
    c.alignment = BANNER_META_ALIGN
    ws.row_dimensions[generated_row].height = BANNER_GENERATED_ROW_HEIGHT

    return generated_row + 1


def _write_metadata_grid(ws: Worksheet, start_row: int, pairs: list[tuple[str, Any]]) -> int:
    row = start_row
    for label, value in pairs:
        lc = ws.cell(row=row, column=1, value=label)
        lc.font = META_LABEL_FONT
        lc.fill = META_FILL
        lc.border = THIN_BORDER
        lc.alignment = LEFT

        vc = ws.cell(row=row, column=2, value=_clean_value(value))
        vc.font = META_VALUE_FONT
        vc.border = THIN_BORDER
        vc.alignment = LEFT
        row += 1
    return row


def _write_section_heading(ws: Worksheet, row: int, text: str, *, col_span: int) -> int:
    for col in range(1, col_span + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = SECTION_FILL
        cell.font = SECTION_FONT
        cell.border = THIN_BORDER
        cell.alignment = LEFT if col == 1 else CENTER
    ws.cell(row=row, column=1, value=text)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_span)
    return row + 1


def _write_formatted_table(
    ws: Worksheet,
    start_row: int,
    headers: list[str],
    rows: list[list[Any]],
    *,
    column_formats: list[str],
    freeze: bool,
    table_name: str,
    wrap_columns: set[int] | None = None,
) -> TableBounds:
    wrap_columns = wrap_columns or set()
    col_count = len(headers)
    header_row = start_row

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    data_start = start_row + 1
    for offset, row_values in enumerate(rows):
        excel_row = data_start + offset
        stripe = ALT_FILL if offset % 2 else None
        for col_idx, value in enumerate(row_values, 1):
            cell = ws.cell(row=excel_row, column=col_idx, value=value)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            if stripe:
                cell.fill = stripe

            fmt = column_formats[col_idx - 1] if col_idx - 1 < len(column_formats) else GENERAL_FMT
            if fmt == PCT_FMT and isinstance(cell.value, (int, float)):
                cell.value = _excel_percent(cell.value)
            elif fmt == CURRENCY_FMT and isinstance(cell.value, (int, float)):
                cell.value = float(cell.value)
            elif fmt == COUNT_FMT and isinstance(cell.value, (int, float)):
                cell.value = int(float(cell.value))
            elif fmt == SCORE_FMT and isinstance(cell.value, (int, float)):
                cell.value = float(cell.value)
            if isinstance(cell.value, str) and cell.value != "":
                cell.number_format = GENERAL_FMT
            else:
                cell.number_format = fmt

            if col_idx in wrap_columns or fmt == GENERAL_FMT and isinstance(value, str) and len(str(value)) > 40:
                cell.alignment = TOP_ALIGN
            elif fmt in (CURRENCY_FMT, PCT_FMT, COUNT_FMT, SCORE_FMT):
                cell.alignment = RIGHT
            else:
                cell.alignment = LEFT

    last_row = data_start + len(rows) - 1 if rows else start_row
    last_col = col_count
    ref = f"A{header_row}:{get_column_letter(last_col)}{max(last_row, header_row)}"
    ws.auto_filter.ref = ref
    if freeze:
        # Freeze banner + column headers while scrolling data
        ws.freeze_panes = ws.cell(row=data_start, column=1)

    return TableBounds(
        header_row=header_row,
        data_start=data_start,
        last_row=max(last_row, header_row),
        last_col=last_col,
        headers=headers,
    )


# ── DataFrame builders ────────────────────────────────────────────────────────


def _products_dataframe(records: list[dict]) -> pd.DataFrame:
    if not records:
        return pd.DataFrame()
    prod_df = pd.DataFrame(records)
    rename = {
        "sku": "SKU",
        "title": "Product",
        "category": "Category",
        "revenue": "Revenue",
        "units_sold": "Units Sold",
        "margin_pct": "Margin %",
        "health_score": "Health Score",
        "performance_tier": "Tier",
        "performance_rank": "Rank",
        "price": "Price",
        "cost": "Cost",
    }
    if "trend" in prod_df.columns:
        rename["trend"] = "Trend"
    elif "trend_indicator" in prod_df.columns:
        rename["trend_indicator"] = "Trend"
    prod_df = prod_df.rename(columns={k: v for k, v in rename.items() if k in prod_df.columns})
    prod_df = prod_df.loc[:, ~prod_df.columns.duplicated()]
    preferred = [
        "SKU",
        "Product",
        "Category",
        "Tier",
        "Revenue",
        "Units Sold",
        "Margin %",
        "Health Score",
        "Rank",
        "Trend",
        "Price",
        "Cost",
    ]
    cols = [c for c in preferred if c in prod_df.columns]
    extra = [c for c in prod_df.columns if c not in cols]
    return prod_df[cols + extra]


def _inventory_dataframe(analysis: dict, inventory_rows: list[dict]) -> pd.DataFrame:
    inv_risk = analysis.get("inventory_risk", {})
    inv_records: list[dict] = []
    for alert in inv_risk.get("alerts", []):
        inv_records.append({
            "SKU": alert.get("sku"),
            "Risk Type": (alert.get("type") or "").replace("_", " ").title(),
            "Severity": alert.get("severity"),
            "Quantity": alert.get("quantity"),
            "Message": alert.get("message"),
            "Recommendation": alert.get("recommendation"),
        })
    for suggestion in inv_risk.get("reorder_suggestions", []):
        inv_records.append({
            "SKU": suggestion.get("sku"),
            "Risk Type": "Reorder Suggestion",
            "Days of Cover": suggestion.get("days_of_cover"),
            "Suggested Reorder": suggestion.get("suggested_reorder"),
            "Urgency": suggestion.get("urgency"),
        })

    if inventory_rows:
        base = pd.DataFrame(inventory_rows)
        rename = {
            "sku": "SKU",
            "quantity_on_hand": "Qty On Hand",
            "inventory_health_score": "Health Score",
            "risk_level": "Risk Level",
            "days_in_stock": "Days In Stock",
            "days_cover": "Days of Cover",
            "inventory_risk": "Risk Level",
        }
        base = base.rename(columns={k: v for k, v in rename.items() if k in base.columns})
        if inv_records:
            extra = pd.DataFrame(inv_records)
            base = base.loc[:, ~base.columns.duplicated()]
            extra = extra.loc[:, ~extra.columns.duplicated()]
            return pd.concat([base, extra], ignore_index=True, sort=False)
        return base.loc[:, ~base.columns.duplicated()]
    return pd.DataFrame(inv_records)


def _alerts_dataframe(alerts_rows: list[dict]) -> pd.DataFrame:
    alerts_df = pd.DataFrame(alerts_rows)
    if alerts_df.empty:
        return alerts_df
    if "entity_id" in alerts_df.columns and "sku" not in alerts_df.columns:
        alerts_df = alerts_df.rename(columns={"entity_id": "sku"})
    return alerts_df.rename(
        columns={
            "severity": "Severity",
            "alert_type": "Issue Type",
            "message": "Message",
            "created_at": "Timestamp",
            "sku": "SKU",
            "title": "Title",
        }
    )


def _profit_leakage_dataframe(analysis: dict) -> pd.DataFrame:
    leakage = analysis.get("profit_leakage", {}).get("issues", [])
    leak_df = pd.DataFrame(leakage)
    if leak_df.empty:
        return leak_df
    return leak_df.rename(
        columns={
            "sku": "SKU",
            "title": "Product",
            "type": "Issue",
            "estimated_impact": "Estimated Impact",
            "message": "Details",
            "recommendation": "Recommendation",
            "severity": "Severity",
        }
    )


# ── Formatting helpers ────────────────────────────────────────────────────────


def _apply_typed_value(cell: Any, value: Any, kind: str) -> None:
    """Write a KPI cell with explicit Excel number format (never infer from column index)."""
    if value is None or value == "":
        cell.value = "—"
        cell.number_format = GENERAL_FMT
        return

    try:
        if kind in ("currency", "exposure_currency"):
            cell.value = abs(float(value))
            cell.number_format = CURRENCY_FMT
        elif kind == "percent":
            cell.value = _excel_percent(value)
            cell.number_format = PCT_FMT
        elif kind == "count":
            cell.value = int(float(value))
            cell.number_format = COUNT_FMT
        elif kind == "score":
            cell.value = float(value)
            cell.number_format = SCORE_FMT
        else:
            cell.value = _clean_value(value)
            cell.number_format = GENERAL_FMT
    except (TypeError, ValueError):
        cell.value = _clean_value(value)
        cell.number_format = GENERAL_FMT


def _value_format_for_kind(kind: str) -> str:
    if kind in ("currency", "exposure_currency"):
        return CURRENCY_FMT
    if kind == "percent":
        return PCT_FMT
    if kind == "count":
        return COUNT_FMT
    if kind == "score":
        return SCORE_FMT
    return GENERAL_FMT


_CURRENCY_HEADERS = frozenset({
    "revenue",
    "price",
    "cost",
    "estimated impact",
    "total revenue",
    "dead inventory value",
    "profit leakage estimate",
    "avg order value",
    "inventory value",
})


def _column_format(name: str) -> str:
    lower = name.lower().strip()

    if lower in _CURRENCY_HEADERS or lower.endswith(" revenue"):
        return CURRENCY_FMT
    if "estimated impact" in lower or "order value" in lower:
        return CURRENCY_FMT

    if (
        "margin" in lower
        or lower.endswith("_pct")
        or lower.endswith("%")
        or "efficiency" in lower
    ) and "score" not in lower:
        return PCT_FMT

    if "health score" in lower or lower == "health score":
        return SCORE_FMT

    if lower in {"quantity", "units sold", "count", "orders", "row_count", "active products", "qty on hand"}:
        return COUNT_FMT

    if "score" in lower or "rank" in lower or "days" in lower:
        if "revenue" in lower:
            return CURRENCY_FMT
        if "score" in lower:
            return SCORE_FMT
        return COUNT_FMT

    if lower in {"severity", "issue type", "trend", "tier", "category", "sku", "product", "message", "details"}:
        return GENERAL_FMT

    return GENERAL_FMT


def _should_wrap(name: str) -> bool:
    lower = name.lower()
    return any(k in lower for k in ("message", "details", "recommendation", "notes", "formula"))


def _excel_percent(value: Any) -> Any:
    if value is None or value == "":
        return None
    try:
        v = float(value)
    except (TypeError, ValueError):
        return value
    if abs(v) > 1.5:
        return v / 100.0
    return v


def _trace_notes(trace: dict) -> str:
    parts: list[str] = []
    if trace.get("notes"):
        parts.append(str(trace["notes"]))
    missing = trace.get("missing_columns") or []
    if missing:
        parts.append(f"Missing columns: {', '.join(missing)}")
    cols = trace.get("columns_used") or []
    if cols:
        parts.append(f"Columns used: {', '.join(cols)}")
    return "; ".join(parts)


def _format_datetime(value: Any) -> str:
    if hasattr(value, "strftime"):
        return format_display(value)
    return str(value)


def _format_count(value: Any) -> str:
    if value is None:
        return "—"
    try:
        return f"{int(value):,}"
    except (TypeError, ValueError):
        return str(value)


def _set_column_widths(ws: Worksheet, widths: dict[int, float]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width


def _auto_width(
    ws: Worksheet,
    *,
    headers: list[str] | None = None,
    min_width: float = 12,
    max_width: float = 52,
) -> None:
    wrap_max = 56
    for col_cells in ws.columns:
        col_idx = col_cells[0].column
        letter = get_column_letter(col_idx)
        header_name = ""
        if headers and col_idx <= len(headers):
            header_name = headers[col_idx - 1]
        cap = wrap_max if _should_wrap(header_name) else max_width
        max_len = 0
        for cell in col_cells:
            if cell.value is None:
                continue
            max_len = max(max_len, min(len(str(cell.value)), 80))
        ws.column_dimensions[letter].width = min(max(max_len + 2, min_width), cap)


def _clean_value(value: Any) -> Any:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    if hasattr(value, "isoformat"):
        return value.isoformat(sep=" ", timespec="seconds")
    return value
