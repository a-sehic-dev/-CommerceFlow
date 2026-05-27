"""Enterprise presentation polish: conditional formatting, charts, tab colors, protection."""

from __future__ import annotations

from dataclasses import dataclass
from typing import Any

from openpyxl.drawing.image import Image as XLImage
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# Tab colors (RGB hex without #)
SHEET_TAB_COLORS: dict[str, str] = {
    "Executive Summary": "1E3A8A",
    "Products Intelligence": "6D28D9",
    "Inventory Health": "D97706",
    "Alerts Center": "DC2626",
    "Profit Leakage": "7F1D1D",
}

# Conditional formatting palette
CF_CRITICAL_FILL = PatternFill("solid", fgColor="991B1B")
CF_CRITICAL_FONT = Font(color="FFFFFF", bold=True)
CF_HIGH_FILL = PatternFill("solid", fgColor="EA580C")
CF_HIGH_FONT = Font(color="FFFFFF", bold=True)
CF_MEDIUM_FILL = PatternFill("solid", fgColor="FDE047")
CF_MEDIUM_FONT = Font(color="713F12", bold=True)
CF_LOW_FILL = PatternFill("solid", fgColor="E2E8F0")
CF_LOW_FONT = Font(color="334155", bold=True)

CF_GREEN_FONT = Font(color="15803D", bold=True)
CF_RED_FONT = Font(color="B91C1C", bold=True)
CF_ORANGE_FONT = Font(color="C2410C", bold=True)

CF_HEALTH_HIGH = PatternFill("solid", fgColor="DCFCE7")
CF_HEALTH_MID = PatternFill("solid", fgColor="FEF3C7")
CF_HEALTH_LOW = PatternFill("solid", fgColor="FEE2E2")

EXEC_COLS = 12

CHART_PALETTE = (
    "1E3A8A",
    "4F46E5",
    "0EA5E9",
    "059669",
    "D97706",
    "DC2626",
    "7C3AED",
    "64748B",
)

CHART_NAVY = "1E3A8A"
CHART_INDIGO = "4F46E5"
CHART_SKY = "0EA5E9"
CHART_GREEN = "059669"
CHART_AMBER = "D97706"
CHART_RED = "DC2626"

INVENTORY_RISK_COLORS = {
    "Low": CHART_GREEN,
    "Medium": CHART_AMBER,
    "Critical": CHART_RED,
}

CHART_DATA_START_COL = 14
CHART_PANEL_FILL = PatternFill("solid", fgColor="F1F5F9")
CHART_PANEL_BORDER = Border(
    left=Side(style="thin", color="E2E8F0"),
    right=Side(style="thin", color="E2E8F0"),
    top=Side(style="thin", color="E2E8F0"),
    bottom=Side(style="thin", color="E2E8F0"),
)


@dataclass
class TableBounds:
    header_row: int
    data_start: int
    last_row: int
    last_col: int
    headers: list[str]

    def col_index(self, name: str) -> int | None:
        target = name.lower().strip()
        for idx, h in enumerate(self.headers, 1):
            if h.lower().strip() == target:
                return idx
        return None

    def col_letter(self, name: str) -> str | None:
        idx = self.col_index(name)
        return get_column_letter(idx) if idx else None

    def data_range(self, col_name: str) -> str | None:
        letter = self.col_letter(col_name)
        if not letter or self.last_row < self.data_start:
            return None
        return f"{letter}{self.data_start}:{letter}{self.last_row}"


def apply_workbook_tab_colors(wb) -> None:
    for ws in wb.worksheets:
        color = SHEET_TAB_COLORS.get(ws.title)
        if color:
            ws.sheet_properties.tabColor = color


def apply_severity_column_rules(ws: Worksheet, bounds: TableBounds, *, col_name: str = "Severity") -> None:
    rng = bounds.data_range(col_name)
    if not rng:
        return
    letter = bounds.col_letter(col_name)
    anchor = f"${letter}{bounds.data_start}"
    rules = [
        (f'LOWER({anchor})="critical"', CF_CRITICAL_FILL, CF_CRITICAL_FONT),
        (f'LOWER({anchor})="high"', CF_HIGH_FILL, CF_HIGH_FONT),
        (f'LOWER({anchor})="medium"', CF_MEDIUM_FILL, CF_MEDIUM_FONT),
        (f'LOWER({anchor})="low"', CF_LOW_FILL, CF_LOW_FONT),
    ]
    for formula, fill, font in rules:
        ws.conditional_formatting.add(
            rng,
            FormulaRule(formula=[formula], fill=fill, font=font),
        )


def apply_trend_column_rules(ws: Worksheet, bounds: TableBounds, *, col_name: str = "Trend") -> None:
    rng = bounds.data_range(col_name)
    if not rng:
        return
    letter = bounds.col_letter(col_name)
    anchor = f"${letter}{bounds.data_start}"
    ws.conditional_formatting.add(
        rng,
        FormulaRule(
            formula=[f'OR(ISNUMBER(SEARCH("rising",{anchor})),ISNUMBER(SEARCH("↑",{anchor})))'],
            font=CF_GREEN_FONT,
        ),
    )
    ws.conditional_formatting.add(
        rng,
        FormulaRule(
            formula=[f'OR(ISNUMBER(SEARCH("declining",{anchor})),ISNUMBER(SEARCH("↓",{anchor})))'],
            font=CF_RED_FONT,
        ),
    )
    ws.conditional_formatting.add(
        rng,
        FormulaRule(
            formula=[f'OR(ISNUMBER(SEARCH("unstable",{anchor})),ISNUMBER(SEARCH("volatile",{anchor})))'],
            font=CF_ORANGE_FONT,
        ),
    )


def apply_health_score_rules(ws: Worksheet, bounds: TableBounds, *, col_name: str = "Health Score") -> None:
    rng = bounds.data_range(col_name)
    if not rng:
        return
    ws.conditional_formatting.add(rng, CellIsRule(operator="greaterThanOrEqual", formula=["80"], fill=CF_HEALTH_HIGH))
    ws.conditional_formatting.add(
        rng,
        CellIsRule(operator="between", formula=["50", "79.99"], fill=CF_HEALTH_MID),
    )
    ws.conditional_formatting.add(rng, CellIsRule(operator="lessThan", formula=["50"], fill=CF_HEALTH_LOW))


def apply_risk_level_rules(ws: Worksheet, bounds: TableBounds, *, col_name: str = "Risk Level") -> None:
    rng = bounds.data_range(col_name)
    if not rng:
        return
    letter = bounds.col_letter(col_name)
    anchor = f"${letter}{bounds.data_start}"
    ws.conditional_formatting.add(
        rng,
        FormulaRule(formula=[f'LOWER({anchor})="high"'], fill=CF_CRITICAL_FILL, font=CF_CRITICAL_FONT),
    )
    ws.conditional_formatting.add(
        rng,
        FormulaRule(formula=[f'LOWER({anchor})="medium"'], fill=CF_MEDIUM_FILL, font=CF_MEDIUM_FONT),
    )
    ws.conditional_formatting.add(
        rng,
        FormulaRule(formula=[f'LOWER({anchor})="low"'], fill=CF_HEALTH_HIGH, font=CF_GREEN_FONT),
    )


def apply_sheet_conditional_formatting(ws: Worksheet, bounds: TableBounds, sheet_name: str) -> None:
    if sheet_name == "Alerts Center":
        apply_severity_column_rules(ws, bounds)
    elif sheet_name == "Profit Leakage":
        apply_severity_column_rules(ws, bounds)
    elif sheet_name == "Products Intelligence":
        apply_trend_column_rules(ws, bounds)
        apply_health_score_rules(ws, bounds)
    elif sheet_name == "Inventory Health":
        apply_risk_level_rules(ws, bounds)
        if bounds.col_index("Severity"):
            apply_severity_column_rules(ws, bounds)
        if bounds.col_index("Urgency"):
            apply_severity_column_rules(ws, bounds, col_name="Urgency")


def protect_executive_sheet(ws: Worksheet, *, unlock_from_row: int) -> None:
    """Light protection on audit log only — charts must remain fully interactive."""
    from openpyxl.styles import Protection

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=EXEC_COLS):
        for cell in row:
            cell.protection = Protection(locked=False)

    ws.protection.sheet = False
    ws.protection.enableObjects = True
    ws.protection.enableScenarios = False


def _paint_chart_panel(ws: Worksheet, top_row: int, bottom_row: int) -> None:
    """Soft shaded backdrop behind executive chart images."""
    for r in range(top_row, bottom_row + 1):
        ws.row_dimensions[r].height = max(ws.row_dimensions[r].height or 0, 28)
        for c in range(1, EXEC_COLS + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = CHART_PANEL_FILL
            cell.border = CHART_PANEL_BORDER


def _embed_chart_png(ws: Worksheet, anchor: str, png_bytes: bytes, *, width: int, height: int) -> None:
    """Embed a rendered chart image — visible in Excel even in Protected View."""
    import tempfile
    from io import BytesIO
    from pathlib import Path

    img = None
    tmp_path: Path | None = None
    try:
        img = XLImage(BytesIO(png_bytes))
    except Exception:
        with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp:
            tmp.write(png_bytes)
            tmp.flush()
            tmp_path = Path(tmp.name)
        img = XLImage(str(tmp_path))
    try:
        img.width = width
        img.height = height
        ws.add_image(img, anchor)
    finally:
        if tmp_path and tmp_path.exists():
            tmp_path.unlink(missing_ok=True)


def _write_chart_table(
    ws: Worksheet,
    anchor_row: int,
    label_col: int,
    value_col: int,
    headers: tuple[str, str],
    rows: list[tuple[Any, Any]],
) -> None:
    """Mirror chart series into worksheet cells (audit + Excel recovery)."""
    ws.cell(row=anchor_row, column=label_col, value=headers[0])
    ws.cell(row=anchor_row, column=value_col, value=headers[1])
    for i, (label, value) in enumerate(rows):
        r = anchor_row + 1 + i
        ws.cell(row=r, column=label_col, value=label)
        ws.cell(row=r, column=value_col, value=float(value) if isinstance(value, (int, float)) else value)


def _prepare_revenue_series(trend: list[dict]) -> tuple[list[str], list[float]]:
    rows = [
        (
            str(p.get("date") or p.get("period") or ""),
            float(p.get("revenue") or 0),
        )
        for p in trend[:90]
    ]
    if not rows or sum(v for _, v in rows) <= 0:
        return [], []
    labels = [r[0] for r in rows]
    values = [r[1] for r in rows]
    return labels, values


def _prepare_category_series(categories: list[dict]) -> tuple[list[str], list[float]]:
    rows = [
        (str(c.get("category") or "Uncategorized"), float(c.get("revenue") or 0))
        for c in categories
        if float(c.get("revenue") or 0) > 0
    ]
    if not rows:
        return [], []
    labels = [r[0] for r in rows]
    values = [r[1] for r in rows]
    return labels, values


def _prepare_inventory_series(risk_counts: dict[str, int]) -> tuple[list[str], list[float]]:
    order = ("Low", "Medium", "Critical")
    values = [float(int(risk_counts.get(label, 0))) for label in order]
    if sum(values) <= 0:
        return [], []
    return list(order), values


def validate_executive_charts(ws: Worksheet) -> None:
    """Log when chart images are missing; never block workbook export."""
    import logging

    images = getattr(ws, "_images", None) or []
    if len(images) < 3:
        logging.getLogger("commerceflow.export").info(
            "Executive Summary chart images: %s of 3 embedded (data tables still included)",
            len(images),
        )


def _try_embed_chart(
    ws: Worksheet,
    anchor: str,
    png_bytes: bytes,
    *,
    width: int,
    height: int,
    chart_name: str = "chart",
) -> bool:
    if not png_bytes:
        import logging

        logging.getLogger("commerceflow.export").warning(
            "Chart not embedded (%s): empty render output", chart_name
        )
        return False
    try:
        _embed_chart_png(ws, anchor, png_bytes, width=width, height=height)
        return True
    except Exception as exc:
        import logging

        logging.getLogger("commerceflow.export").warning(
            "Chart not embedded (%s): %s", chart_name, exc
        )
        return False


def write_executive_charts(ws: Worksheet, start_row: int, chart_data: dict[str, Any]) -> int:
    """Render executive chart images when matplotlib is available; always write data tables."""
    from openpyxl.styles import Alignment, Font, PatternFill

    from app.utils.chart_images import (
        charts_available,
        render_category_doughnut_png,
        render_inventory_bar_png,
        render_revenue_line_png,
    )

    section_fill = PatternFill("solid", fgColor="0F172A")
    section_font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")

    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=EXEC_COLS)
    bar = ws.cell(row=start_row, column=1, value="PERFORMANCE INSIGHTS")
    bar.fill = section_fill
    bar.font = section_font
    bar.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=2)
    ws.row_dimensions[start_row].height = 30

    chart_row = start_row + 2
    chart_panel_bottom = chart_row + 11
    ws.row_dimensions[start_row + 1].height = 8
    ws.row_dimensions[chart_row].height = 28
    _paint_chart_panel(ws, chart_row, chart_panel_bottom)

    trend = chart_data.get("revenue_trend") or []
    categories = chart_data.get("category_breakdown") or []
    inventory_risk = chart_data.get("inventory_risk") or {}

    data_base_row = chart_row + 20

    rev_labels, rev_values = _prepare_revenue_series(trend)
    _write_chart_table(
        ws, data_base_row, 1, 2, ("Date", "Revenue"), list(zip(rev_labels, rev_values, strict=True))
    )
    embedded: list[bool] = []
    embedded.append(
        _try_embed_chart(
            ws,
            f"B{chart_row}",
            render_revenue_line_png(rev_labels, rev_values),
            width=490,
            height=268,
            chart_name="Revenue Trend",
        )
    )

    cat_labels, cat_values = _prepare_category_series(categories)
    _write_chart_table(
        ws,
        data_base_row,
        4,
        5,
        ("Category", "Revenue"),
        list(zip(cat_labels, cat_values, strict=True)),
    )
    embedded.append(
        _try_embed_chart(
            ws,
            f"F{chart_row}",
            render_category_doughnut_png(cat_labels, cat_values),
            width=470,
            height=268,
            chart_name="Category Revenue Mix",
        )
    )

    inv_labels, inv_values = _prepare_inventory_series(inventory_risk)
    _write_chart_table(
        ws,
        data_base_row,
        7,
        8,
        ("Risk Level", "SKUs"),
        list(zip(inv_labels, inv_values, strict=True)),
    )
    embedded.append(
        _try_embed_chart(
            ws,
            f"J{chart_row}",
            render_inventory_bar_png(inv_labels, inv_values),
            width=450,
            height=268,
            chart_name="Inventory Risk Breakdown",
        )
    )

    if not any(embedded):
        import logging

        note_row = chart_row + 1
        ws.merge_cells(start_row=note_row, start_column=2, end_row=note_row, end_column=EXEC_COLS - 1)
        if not charts_available():
            note_text = (
                "Chart images require matplotlib in the server environment. "
                "Install requirements.txt and restart CommerceFlow. KPI and audit tables are included below."
            )
        else:
            note_text = (
                "Chart rendering was skipped for this export. KPI and audit tables are included below."
            )
        note = ws.cell(row=note_row, column=2, value=note_text)
        note.font = Font(name="Calibri", size=9, italic=True, color="64748B")
        note.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        logging.getLogger("commerceflow.export").warning(
            "Executive Summary: 0/3 chart images embedded"
        )

    validate_executive_charts(ws)

    return chart_panel_bottom + 2
